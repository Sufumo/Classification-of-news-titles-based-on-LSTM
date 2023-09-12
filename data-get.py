import time
import requests
import json
import openpyxl as op


def write_xlsx(data, path, sheet_name):
    wb = op.load_workbook(path)
    ws = None
    try:
        ws = wb[sheet_name]
    except:
        wb.create_sheet(sheet_name)
        ws = wb[sheet_name]
    finally:
        write_point = ws.max_row + 1
        for i, j in enumerate(data.values()):
            data_type = type(j)
            if data_type == str or data_type == int or data_type == float:
                ws.cell(row=write_point, column=i + 1).value = j
        wb.save(filename=path)


user = {
      "cookie": r"buvid3=9CC34503-021B-B5C0-700D-05095A77BC1093230infoc; i-wanna-go-back=-1; _uuid=39989372-95C8-898E-1D"
                r"EE-BC433B9C2C4803896infoc; FEED_LIVE_VERSION=V8; home_feed_column=5; CURRENT_FNVAL=4048; DedeUserID=51"
                r"7408168; DedeUserID__ckMd5=ed2c6d95e961bb10; b_ut=5; header_theme_version=CLOSE; nostalgia_conf=-1; CU"
                r"RRENT_PID=9125c390-df12-11ed-890c-69a78663bcc9; rpdid=|(J|J|Rk)JJ|0J'uY)uml|Ykk; browser_resolution=15"
                r"00-889; buvid_fp_plain=undefined; b_nut=1683388017; SESSDATA=3cd98f37%2C1700116558%2Ca723a%2A52; bili_"
                r"jct=3d35621a9caf01cd3eef0d03eb16423a; fingerprint=baea25c2fc4df1760f2d995c5683be08; buvid_fp=0f998ba4d"
                r"0428366aef8ad400ffe61d9; sid=7jvich79; buvid4=EC529DE1-FEC1-7179-612C-C8AB96EB9DE996881-023042000-FWSn"
                r"vHx5X1sbxS/9z+GctQ%3D%3D; b_lsid=96C99F49_189354D80D2; PVID=2; bp_video_offset_517408168=8159505854591"
                r"14100",
      'referer': r'https://space.bilibili.com/517408168/favlist?spm_id_from=333.999.0.0',
      "user-agent": r"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 " \
                    r"Safari/537.36"
}
page_num = 44
for page in range(1, page_num+1):
    url = f"https://api.bilibili.com/x/v3/fav/resource/list?media_id=914685768&pn={page}&" \
          f"ps=20&keyword=&order=mtime&type=0&tid=0&platform=web"
    res1 = requests.get(url=url, headers=user)
    print(res1.text)
    data1 = json.loads(res1.text)
    print(data1)
    try:
        iter1 = data1["data"]["medias"]
        for i in iter1:
            write_xlsx(data=i, path="./test.xlsx", sheet_name="data")
        print(f"单元{page}写入成功")
    except Exception as e:
        print(f"单元{page}写入错误,错误代码为：", e)
    finally:
        time.sleep(1)

