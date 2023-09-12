import openpyxl as op
import pandas as pd
import torch


def write_xlsx(data, path, sheet_name):
     wb = op.load_workbook(path)
     ws = None
     try:
          ws = wb[sheet_name]
     except:
          wb.create_sheet(sheet_name)
          ws = wb[sheet_name]
     finally:
          write_point= ws.max_row + 1
          for i,j in enumerate(data.values()):
               ws.cell(row=write_point, column=i+1).value = j
          wb.save(filename=path)


if __name__ == '__main__':
     df = pd.read_csv("info.csv")
     print(df["category"].value_counts())
     # a = torch.tensor([1,2,3,4,5])
     # max_list = []
     # # print(a.argmax().item())
     # for i in range(3):
     #      max_list.append(a.argmax().item())
     #      a[a.argmax().item()] = torch.tensor([0])
     # print(max_list)
     # a = []
     # for i in range(1, 191345):
     #      if 191344 % i == 0:
     #           a.append(i)
     # a.sort(key=int)
     # print(a)
     # t1 = torch.tensor([[0.1,0,0],[0,0.1,0],[0,0,0.1]])
     # # print(t1.squeeze(0))
     # # # print(t1.shape)
     # # # t1 = torch.randn(3,3)
     # # # print(t1.shape)
     # t2 = torch.tensor([0,1,2])
     # # print(t1)
     # # print(t2)
     # t3 = torch.nn.CrossEntropyLoss()
     # print(torch.eq(torch.max(t1, dim=1)[1],t2))
     # print(t2.argmax())
     # print(torch.Tensor([1,2,3]).argmax().item())
    # df = pd.read_csv("./info.csv")
    # print(df["category"].value_counts())
    # a = {"1":1,"2":2}
    # path = "./test.xlsx"
    # sheet_name = "data"
    # write_xlsx(a, path, sheet_name)
#     str = '''100 民生 故事 news_story
# 101 文化 文化 news_culture
# 102 娱乐 娱乐 news_entertainment
# 103 体育 体育 news_sports
# 104 财经 财经 news_finance
# 106 房产 房产 news_house
# 107 汽车 汽车 news_car
# 108 教育 教育 news_edu
# 109 科技 科技 news_tech
# 110 军事 军事 news_military
# 112 旅游 旅游 news_travel
# 113 国际 国际 news_world
# 114 证券 股票 stock
# 115 农业 三农 news_agriculture
# 116 电竞 游戏 news_game'''
#     li1 = []
#     for i in str.split("\n"):
#         j = i.split()
#         li1.append(f"{j[1]} {j[2]}")
#     print(li1)

