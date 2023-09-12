import pandas as pd
import torchtext
import pkuseg
import torch
from torch.utils.data import Dataset, DataLoader
import functools
import json
import time
import openpyxl as op
seg = pkuseg.pkuseg()


class DataFrameDataset(Dataset):
    def __init__(self, content, label):
        self.content = content
        self.label = label

    def __getitem__(self, index):
        return self.content[index], self.label[index]

    def __len__(self):
        return len(self.label)


def collate_batch(data):
    unit_x = []
    unit_y = []
    for unit in data:
        unit_x.append(unit[0])
        unit_y.append(unit[1])
    return {"text": unit_x,
            "label": torch.tensor(unit_y)}


def yield_tokens(data_df):
    for _, row in data_df.iterrows():
        yield [token for token in seg.cut(row.title)]


def write_xlsx(data, path, sheet_name):
    wb = op.load_workbook(path)
    ws = None
    try:
        ws = wb[sheet_name]
    except:
        wb.create_sheet(sheet_name)
        ws = wb[sheet_name]
    finally:
        write_point = ws.max_row + 1
        for i, j in enumerate(data):
            data_type = type(j)
            if data_type == str or data_type == int or data_type == float:
                ws.cell(row=write_point+i, column=1).value = j
        wb.save(filename=path)


if __name__ == '__main__':
    # emb = torch.nn.Embedding(100, 2)
    # b1 = emb(torch.tensor([[0, 1, 2],
    #                       [0, 6, 7]]))
    # print(b1.shape)
    # print(b1.permute(1,0,2))
    # seg = pkuseg.pkuseg()
    df = pd.read_csv("./info.csv")
    list2 = []
    for i in df.index:
        list2.append(seg.cut(df.loc[i, "title"]))
    list1 =[len(x) for x in list2]
    write_xlsx(data=list1, path="./data_analysis.xlsx", sheet_name="data")
    print(max(list1))
    print(sum(list1)/len(list1))
    # dic1 = {"category":[], "title":[]}
    # for i in df.index:
    #     dic1["category"].append(df.loc[i, "category"])
    #     dic1["title"].append(df.loc[i, "title"])
    # cs2 = pd.DataFrame(dic1)
    # cs2.to_csv("info.csv")
    # train_iter = DataFrameDataset(list(df['title']), list(df['category']))
    # train_loader = DataLoader(train_iter, batch_size=8, shuffle=True
    #                           ,collate_fn=collate_batch)
    # for batch in train_loader:
    #     print(batch)
    #     break
    # vocab = torchtext.vocab.build_vocab_from_iterator(yield_tokens(df), min_freq=10,
    #                                                   max_tokens=20000, specials=['<unk>', '<pad>'])
    # print(len(vocab.get_stoi()))
    # print(vocab.get_itos()[0:10])
    # print(vocab.get_stoi()["<pad>"])
    # int_to_str = {}
    # for i,j in vocab.get_stoi().items():
    #     int_to_str[f"{j}"] = i
    # with open("str_to_int_vocab.json", "w", encoding="utf-8") as f:
    #     f.write(json.dumps(vocab.get_stoi(), ensure_ascii=False))
    # with open("int_to_str_vocab.json", "r", encoding="utf-8") as f:
    #     # f.write(json.dumps(int_to_str, ensure_ascii=False))
    #     vocab = json.loads(f.read())
    #     for i in range(100):
    #         print(vocab[f"{i}"])
    #     js1 = json.loads(f.read())
    #     for i,j in js1.items():
    #         if j == 1000:
    #             print(i,j)
    # iter1 = seg.cut("你好世界,为世界上所有美好而战")
    # for i in iter1:
    #     print(i)
# with open("data_source.txt", "r", encoding="utf-8") as f:
#     dict_keys = ["id", "category", "category_name", "title", "key_words"]
#     news_data = {}
#     for k in dict_keys:
#         news_data[k] = []
#     while f.readline() != "":
#         new_data = f.readline()
#         info = new_data.split("_!_")
#         for i,j in enumerate(info):
#             news_data[dict_keys[i]].append(j)
#     # print(news_data)
#     fc = pd.DataFrame(news_data)
#     print(fc)
#     # fc.to_csv("./info.csv")
# df = pd.read_csv("./info.csv")
# print(df["category"].value_counts())


# for i in df.index:
#     print(df.loc[i, "category"], df.loc[i, "title"])
#     break
